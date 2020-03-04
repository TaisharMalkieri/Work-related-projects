# -*- coding: utf-8 -*-
"""
Created on Mon Jan 13 22:11:50 2020

@author: torong
"""


import os
import datetime as dt
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

def read_information():
    #Read user input
    hist_pos = 'D5'
    save_pos = 'D6'
    
    wb = load_workbook(filename= r'ControlStockRec.xlsx')
    print(wb['Main'][save_pos].value)
    return (wb['Main'][hist_pos].value, hist_pos), (wb['Main'][save_pos].value, save_pos)

def sort_files_by_date(directory_path):
    #Create a dateframe from the directory
    #Files must have the date as title using DD-MM-YYYY as format
    dirpath = os.listdir(directory_path)
    t = pd.DataFrame([[os.path.join(directory_path, y) for y in dirpath],
                       [dt.datetime.strptime(x[:-5], '%d-%m-%Y') for x in dirpath]])
    t = t.transpose()
    t.columns = ['Path', 'Date' ]
    t = t.sort_values('Date', ascending=False)
    t = t.reset_index()
    return t

def create_item_history_dict(file_list):        
    #Item history is a library using item nr as keys. Each item is its own lib
    item_history = {}
    item_status = {}
    
    #For each article that has a diff at day 0, create a dict object in item history
    day0 = pd.read_excel(file_list['Path'].iloc[0])
    day0 = day0.loc[day0['Difference'] != 0]
    for i in range(len(day0['Difference'])):
        item_history[day0['Item number'].iloc[i]] = {
                'Amount': day0['Difference amount'].iloc[i],
                'Difference': day0['Difference'].iloc[i],
                'History': [],
                'DHL': []
                }
        item_status[day0['Item number'].iloc[i]] = day0['Difference'].iloc[i]
    
    active_item_keys = item_status.keys()
    
    
    for f, file in enumerate(file_list['Path']):
        dayX = pd.read_excel(file)
        dayX = dayX.loc[dayX['Difference'] != 0]
        new_active = []
        
        for k, key in enumerate(active_item_keys):
            for x, item in enumerate(dayX['Item number']):
                if key == item:
                    if dayX['Difference'].iloc[x] == item_status[key]:
                        new_active.append(key)
                        pass
                    else:
                        item_history[key]['History'].append((f, item_status[key], file_list['Date'].iloc[f]))
                        item_status[key] = dayX['Difference'].iloc[x]
                        new_active.append(key)
                    
        obsolete = [e for e in active_item_keys if e not in new_active]
        active_item_keys = new_active
        
        for obs in obsolete:
            item_history[obs]['History'].append((f, item_status[obs], str(file_list['Date'].iloc[f])))
        
    for act in active_item_keys:
        item_history[act]['History'].append((len(file_list['Path']), item_status[act], str(file_list['Date'].iloc[-1])))

    return item_history



def write_history(dict_items, dates, save_path):
    wb = Workbook()
    ws_MAIN = wb.active
    
    ws_MAIN.title = 'Main'
    sorted_dict = sorted(dict_items.keys(), key = lambda x: dict_items[x]['Amount'])
    print(ws_MAIN)
    
    
    #ws_MAIN.write(0, 0, 'IndX')
    #ws_MAIN.write(0, 1, 'Comment 0')
    #ws_MAIN.write(0, 2, 'Comment 1')
    #ws_MAIN.write(0 , 3,  'Item')
    #ws_MAIN.write(0 , 4,  'Amount')
    ws_MAIN['A1'] = 'Indx'
    ws_MAIN['B1'] = 'Comment 0'
    ws_MAIN['C1'] = 'Comment 1'
    ws_MAIN['D1'] = 'Item nr'
    ws_MAIN['E1'] = 'Amount'
    
    for d, date in enumerate(dates):
        #ws_MAIN.write(0 , 5+d, str(date)[:-9])
        cl = ws_MAIN.cell(row=1, column=6+d)
        cl.value = str(date)[:-9]
        
    #ws_MAIN.write(0, d+6, 'Unnknown')
    unnknown = ws_MAIN.cell(row=1, column=d+7)
    unnknown.value = 'Unnknown'
    
    for i, key in enumerate(sorted_dict):
        #ws_MAIN.write(i+1, 0, i)
        indx=ws_MAIN.cell(i+2, 1)
        indx.value=i
        
        #ws_MAIN.write(i+1, 3, key)
        item = ws_MAIN.cell(i+2, 4)
        item.value = key
        
        amount = ws_MAIN.cell(i+2, 5)
        amount.value=str(dict_items[key]['Amount'])
    
        diff = ws_MAIN.cell(i+2, 6)
        diff.value=str(dict_items[key]['Difference'])
    
    
        #ws_MAIN.write(i+1, 4, str(dict_items[key]['Amount']))
        #ws_MAIN.write(i+1, 5, str(dict_items[key]['Difference']))
        
        for c, change in enumerate(dict_items[key]['History']):
            incident = ws_MAIN.cell(i+2, 6+change[0])
            incident.value = str(change[1])
            #ws_MAIN.write(i+1, 5+change[0], str(change[1]))
    
    ws_STATS = wb.create_sheet(title='STAT_HIST')
    COUNT_errors = len(dict_items)
    INDX_neg = [x for x in dict_items if dict_items[x]['Difference']<0]
    INDX_pos = [w for w in dict_items if dict_items[w]['Difference']>0]
    COUNT_neg = len(INDX_neg)
    COUNT_pos = COUNT_errors-COUNT_neg
    
    ws_STATS['A1'] = 'Here are the stats of the day'
    ws_STATS['A2'] = 'Count More in AX'
    ws_STATS['A3'] = 'Count More in JDA'
    ws_STATS['A4'] = 'Total '
    
    ws_STATS['A4'] = 'Amount Diff Positiv'
    ws_STATS['A4'] = 'Amount Diff Negativ'
    
    
    today = dt.datetime.today()
    save_path = os.path.join(save_path[0],'Stck_rcn_hist_{}.xlsx'.format(today.strftime("%H_%M_%S")))
    wb.save(save_path)
    return wb
        
                    

hist_path, save_path = read_information()

assert os.path.exists(hist_path[0]),"The path {} at koordinate {} either did not work, or does not work. Spellcheck and assert that you have access to the linked adress".format(hist_path[0], hist_path[1])
assert os.path.exists(save_path[0]),"The path {} at koordinate {} either did not work, or does not work. Spellcheck and assert that you have access to the linked adress".format(save_path[0], save_path[1])

hist_file_paths = sort_files_by_date(hist_path[0])
history_dict = create_item_history_dict(hist_file_paths)
write_history(history_dict, hist_file_paths['Date'], save_path)

